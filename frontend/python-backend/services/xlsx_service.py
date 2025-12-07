from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import tempfile
from typing import List, Dict, Any

class XlsxService:
    """Service for Excel spreadsheet creation and manipulation"""

    @staticmethod
    def create_blank_spreadsheet(title: str, headers: List[str]) -> Path:
        """
        Create a blank spreadsheet with headers.

        Args:
            title: Spreadsheet title
            headers: List of column headers

        Returns:
            Path to created Excel file
        """
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"

        # Title
        sheet['A1'] = title
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells(f'A1:{get_column_letter(len(headers))}1')

        # Headers
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto-size columns
        for idx in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(idx)].width = 15

        temp_dir = Path(tempfile.gettempdir()) / "xlsx-creator"
        temp_dir.mkdir(exist_ok=True)
        output_path = temp_dir / f"blank_{tempfile._get_candidate_names().__next__()}.xlsx"

        wb.save(output_path)
        return output_path

    @staticmethod
    def create_budget_template(year: str, categories: List[str]) -> Path:
        """
        Create a budget spreadsheet template.

        Args:
            year: Budget year
            categories: List of expense categories

        Returns:
            Path to created Excel file
        """
        wb = Workbook()
        sheet = wb.active
        sheet.title = f"{year} Budget"

        # Title
        sheet['A1'] = f"{year} Budget"
        sheet['A1'].font = Font(bold=True, size=16)
        sheet.merge_cells('A1:E1')
        sheet['A1'].alignment = Alignment(horizontal="center")

        # Headers
        headers = ["Category", "Q1", "Q2", "Q3", "Q4", "Total"]
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Categories
        row = 4
        for category in categories:
            sheet.cell(row=row, column=1).value = category
            sheet.cell(row=row, column=1).font = Font(bold=True)

            # Add formulas for totals
            sheet.cell(row=row, column=6).value = f'=SUM(B{row}:E{row})'
            sheet.cell(row=row, column=6).font = Font(bold=True)
            row += 1

        # Totals row
        sheet.cell(row=row, column=1).value = "TOTAL"
        sheet.cell(row=row, column=1).font = Font(bold=True, size=12)
        for col in range(2, 7):
            start_row = 4
            end_row = row - 1
            sheet.cell(row=row, column=col).value = f'=SUM({get_column_letter(col)}{start_row}:{get_column_letter(col)}{end_row})'
            sheet.cell(row=row, column=col).font = Font(bold=True)
            sheet.cell(row=row, column=col).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Column widths
        sheet.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F']:
            sheet.column_dimensions[col].width = 12

        # Number formatting for currency
        for row_num in range(4, row + 1):
            for col in range(2, 7):
                sheet.cell(row=row_num, column=col).number_format = '$#,##0'

        temp_dir = Path(tempfile.gettempdir()) / "xlsx-creator"
        temp_dir.mkdir(exist_ok=True)
        output_path = temp_dir / f"budget_{tempfile._get_candidate_names().__next__()}.xlsx"

        wb.save(output_path)
        return output_path

    @staticmethod
    def create_financial_model(company_name: str, years: List[str]) -> Path:
        """
        Create a basic financial model template.

        Args:
            company_name: Company name
            years: List of years for projection

        Returns:
            Path to created Excel file
        """
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Financial Model"

        # Title
        sheet['A1'] = f"{company_name} - Financial Model"
        sheet['A1'].font = Font(bold=True, size=14)
        sheet.merge_cells(f'A1:{get_column_letter(len(years) + 1)}1')

        # Years headers
        sheet['A3'] = "Line Item"
        sheet['A3'].font = Font(bold=True)
        for idx, year in enumerate(years, start=2):
            cell = sheet.cell(row=3, column=idx)
            cell.value = year
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Financial line items
        line_items = [
            ("ASSUMPTIONS", True),
            ("Revenue Growth %", False),
            ("Gross Margin %", False),
            ("OpEx % of Revenue", False),
            ("", True),
            ("INCOME STATEMENT", True),
            ("Revenue", False),
            ("Cost of Goods Sold", False),
            ("Gross Profit", False),
            ("Operating Expenses", False),
            ("EBITDA", False),
        ]

        row = 4
        for item, is_header in line_items:
            sheet.cell(row=row, column=1).value = item
            if is_header:
                sheet.cell(row=row, column=1).font = Font(bold=True, size=11)
                sheet.cell(row=row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1

        # Add sample formulas
        # Revenue growth
        sheet['B5'].value = 0.15
        sheet['B5'].font = Font(color="0000FF")  # Blue for inputs
        sheet['B5'].number_format = '0.0%'

        # Gross margin
        sheet['B6'].value = 0.60
        sheet['B6'].font = Font(color="0000FF")
        sheet['B6'].number_format = '0.0%'

        # OpEx
        sheet['B7'].value = 0.40
        sheet['B7'].font = Font(color="0000FF")
        sheet['B7'].number_format = '0.0%'

        # Revenue
        sheet['B10'].value = 1000000
        sheet['B10'].font = Font(color="0000FF")
        sheet['B10'].number_format = '$#,##0'

        # Add formulas for future years
        for col_idx in range(3, len(years) + 2):
            col = get_column_letter(col_idx)
            prev_col = get_column_letter(col_idx - 1)

            # Revenue = Prior year * (1 + growth)
            sheet[f'{col}10'].value = f'={prev_col}10*(1+$B$5)'
            sheet[f'{col}10'].number_format = '$#,##0'

            # COGS = Revenue * (1 - Gross Margin)
            sheet[f'{col}11'].value = f'={col}10*(1-$B$6)'
            sheet[f'{col}11'].number_format = '$#,##0'

            # Gross Profit = Revenue - COGS
            sheet[f'{col}12'].value = f'={col}10-{col}11'
            sheet[f'{col}12'].number_format = '$#,##0'

            # OpEx = Revenue * OpEx %
            sheet[f'{col}13'].value = f'={col}10*$B$7'
            sheet[f'{col}13'].number_format = '$#,##0'

            # EBITDA = Gross Profit - OpEx
            sheet[f'{col}14'].value = f'={col}12-{col}13'
            sheet[f'{col}14'].number_format = '$#,##0'

        # First year formulas
        sheet['B11'].value = '=B10*(1-$B$6)'
        sheet['B11'].number_format = '$#,##0'
        sheet['B12'].value = '=B10-B11'
        sheet['B12'].number_format = '$#,##0'
        sheet['B13'].value = '=B10*$B$7'
        sheet['B13'].number_format = '$#,##0'
        sheet['B14'].value = '=B12-B13'
        sheet['B14'].number_format = '$#,##0'

        # Column widths
        sheet.column_dimensions['A'].width = 25
        for col_idx in range(2, len(years) + 2):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 15

        temp_dir = Path(tempfile.gettempdir()) / "xlsx-creator"
        temp_dir.mkdir(exist_ok=True)
        output_path = temp_dir / f"financial_model_{tempfile._get_candidate_names().__next__()}.xlsx"

        wb.save(output_path)
        return output_path

    @staticmethod
    def create_data_table(title: str, data: List[Dict[str, Any]]) -> Path:
        """
        Create a spreadsheet from tabular data.

        Args:
            title: Table title
            data: List of dictionaries with row data

        Returns:
            Path to created Excel file
        """
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Data"

        if not data:
            raise ValueError("Data cannot be empty")

        # Title
        sheet['A1'] = title
        sheet['A1'].font = Font(bold=True, size=14)
        headers = list(data[0].keys())
        sheet.merge_cells(f'A1:{get_column_letter(len(headers))}1')

        # Headers
        for idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=3, column=idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row_data in enumerate(data, start=4):
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(header, "")

        # Auto-size columns
        for idx in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(idx)].width = 15

        temp_dir = Path(tempfile.gettempdir()) / "xlsx-creator"
        temp_dir.mkdir(exist_ok=True)
        output_path = temp_dir / f"data_table_{tempfile._get_candidate_names().__next__()}.xlsx"

        wb.save(output_path)
        return output_path
